import os

import openpyxl as op

'''
Contains Company class with relevant (and maybe some irrelevant) company info

getSortedCompanies() takes in an Excel workbook filename and the Session Name as input, opens the workbook, stores relevant company
information, filters out companies not attending the current session, and then creates a sorted list with the following order:
    Premium Booths, Electricity Booths, Big Companies, then all others (sorted by industry).
'''

class Company:
    def __init__(self, companyName, sessions, industry, boothType, majors, needsElectric, bigComp=False):
        self.companyName = companyName
        self.sessions = sessions
        self.industry = industry
        self.boothType = boothType
        self.majors = majors
        self.needsElectric = needsElectric
        self.bigComp = bigComp
    def printCompanyInfo(self):
        print('-------------------------------------------')
        print('Employer: ', self.companyName)
        print('Sessions: ', self.sessions)
        print('Industry: ', self.industry)
        print('Booth: ', self.boothType)
        print('Majors: ', self.majors)
        print('Needs Electricity?: ', self.needsElectric)
        print('Big Company?: ', self.bigComp)
        
    def clearWrongBooths(self, sessionName):
        booths = self.boothType.split(', ')
        if len(self.sessions.split('; ')) > 1:
            boothCount = 0 # index of the list of booths
            # increment boothCount until the correct index is found (i.e. the one corresponding to the current session)
            for sess in self.sessions.split('; '):
                if sess == sessionName:
                    currSessBooth = booths[boothCount] 
                    break
                else:
                    boothCount += 1
            self.boothType = currSessBooth
        else: # only going to 1 session, so they only need 1 booth- don't need to change anything
            return

def getSortedCompanies(workbookName, sessionName, excludedIndustries, sortBigComps=False):
    wb = op.load_workbook(workbookName)
    sheet = wb.active
    # determine which column number each relevant data is stored in
    for col in range(1, sheet.max_column + 1): 
        currCell = sheet.cell(row=1, column=col).value
        if currCell == 'Employer': employerCol = col
        if currCell == 'Sessions': sessionsCol = col
        if currCell == 'Employer Industry': industryCol = col
        if currCell == 'Requested Booth Options': boothsCol = col
        if currCell == 'Combined Majors': majorsCol = col
        if currCell == 'General Items - Access to Electric': electricCol = col
        if currCell == 'Big Company': bigCompCol = col

    currSessComps = []

    # initialize and store all companies attending the current session
    for row in range(1, sheet.max_row + 1):

        if sessionName in sheet.cell(row=row, column=sessionsCol).value and \
            sheet.cell(row=row, column=industryCol).value not in excludedIndustries:
            
            if sortBigComps:
                newComp = Company(
                    companyName=sheet.cell(row=row, column=employerCol).value, 
                    sessions=sheet.cell(row=row, column=sessionsCol).value,
                    industry=sheet.cell(row=row, column=industryCol).value, 
                    boothType=sheet.cell(row=row, column=boothsCol).value, 
                    majors=sheet.cell(row=row, column=majorsCol).value,
                    needsElectric=bool(sheet.cell(row=row, column=electricCol).value),
                    bigComp=bool(sheet.cell(row=row, column=bigCompCol).value))
            else:
                newComp = Company(
                    companyName=sheet.cell(row=row, column=employerCol).value, 
                    sessions=sheet.cell(row=row, column=sessionsCol).value,
                    industry=sheet.cell(row=row, column=industryCol).value, 
                    boothType=sheet.cell(row=row, column=boothsCol).value, 
                    majors=sheet.cell(row=row, column=majorsCol).value,
                    needsElectric=bool(sheet.cell(row=row, column=electricCol).value),
                    bigComp=False)

            newComp.clearWrongBooths(sessionName)
            currSessComps.append(newComp)
        

    premBooths = []
    elecBooths = []
    bigComps = []
    industries = {}

    # get list of all companies needing premium booths, electric booths, and big companies
    for comp in currSessComps:
        if comp.boothType == 'Premium Booth':
            premBooths.append(comp)
        # store companies that need electric
        if comp.needsElectric:
            elecBooths.append(comp)
        # store companies that are considered 'Big Companies'
        if comp.bigComp:
            bigComps.append(comp)
        # store companies by industry
        if comp.industry not in industries.keys():
            industries[comp.industry] = [comp]
        else:
            industries[comp.industry].append(comp)
    
    # sort current session's companies by premium booths, then electric, then big companies, then add the rest by industry
    sortedComps = []
    for comp in premBooths:
        sortedComps.append(comp)
    for comp in elecBooths:
        if comp not in sortedComps: sortedComps.append(comp)
    for comp in bigComps:
        if comp not in sortedComps: sortedComps.append(comp)
    for compList in industries.values(): # each value in the industries dict is a list of all companies from that industry
        for comp in compList:
            if comp not in sortedComps: sortedComps.append(comp)

    return sortedComps
