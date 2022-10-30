from layout import *
import companies as c
from openpyxl import load_workbook
import os

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


SESSIONNAME = 'Professional Day: Business and Arts and Sciences - Tuesday, Sep 13, 10:00 am - 3:00 pm EDT' #CHANGE THIS LINE!
LAYOUTFILENAME = 'Fall22_CRC_TechDay_Layout.xlsx' #CHANGE THIS LINE!
COMPANIESFILENAME = 'Fall22_All_Registrations.xlsx' #CHANGE THIS LINE!


def main():
    if not os.path.isfile(LAYOUTFILENAME):
        print('\nERROR: LAYOUT NOT FOUND. ENTER THE NAME OF YOUR LAYOUT IN main.py, LINE 8.')
        print('ENSURE THE FILE IS LOCATED IN THE SAME LOCATION AS THIS CODE.')
        return

    if not os.path.isfile(COMPANIESFILENAME):
        print('\nERROR: COMPANIES FILE NOT FOUND. ENTER THE NAME OF YOUR COMAPNIES FILE IN main.py, LINE 9.')
        print('ENSURE THE FILE IS LOCATED IN THE SAME LOCATION AS THIS CODE.')
        return

    wb = load_workbook(filename = LAYOUTFILENAME)
    sheet = wb.worksheets[0]

    excludedIndustries = [] 
    exclInd = input('Enter an industry to be excluded from this layout (Enter nothing to generate layout): ')
    while exclInd != '':
        excludedIndustries = excludedIndustries + [exclInd.title()]
        exclInd = input('Enter an industry to be excluded from this layout (Enter nothing to generate layout): ')

    print('\n\nDo you want to sort by big companies? \nNOTE: This requires a column called Big Company to be manually added')
    print('with a 1 entered for any Big Companies (as determined by the Career Dev Team). All other cells in the column should be left blank.')
    SORT_BIG_COMPS = int(input('Enter 0 for no, or 1 for yes: '))

    startrow = 1
    endrow = len(list(sheet.rows)) 
    startcol = 1 
    endcol = len(list(sheet.columns)) 
    stdBooths = makeBooths(sheet, startrow, endrow, startcol, endcol) # stdBooths is a list of Booth objects

    premBooths = filterPremiumBooths(stdBooths)

    powBooths = filterPowerBooths(stdBooths)

    stdBooths.sort(key=lambda x: x.boothName) 
#finBooths (list of Booth objects) seems to be the list of all the final booths. These booth objects have a booth name AND a company name
    finBooths = []
    comps = c.getSortedCompanies(COMPANIESFILENAME, SESSIONNAME, excludedIndustries, SORT_BIG_COMPS) # comps is a list of Company objects

    if len(comps) > (len(premBooths) + len(powBooths) + len(stdBooths)):
        print('WARNING: NOT ENOUGH BOOTHS IN THE LAYOUT')

    count = 0
    unassignedComps = []

    notEnoughPrem = False
    notEnoughPower = False
    notEnoughBooths = False

    numBoothsAssigned = 0
    for comp in comps:
        if 'Premium Booth' in comp.boothType:
            print(f'{comp.boothType}: {comp.companyName}')
            try:
                pBooth = premBooths[-1] 
                pBooth.companyName = comp.companyName 
                finBooths = finBooths + [pBooth] 
                premBooths = premBooths[0:-1] # premBooths is now equal to exactly what it was before, excluding the last element. 
                                              # this continues until the premBooths has no more elements
                numBoothsAssigned += 1
            except IndexError:
                print(f'WARNING: PREMIUM BOOTH UNABLE TO BE ASSIGNED TO {comp.companyName}')
                notEnoughPrem = True
                unassignedComps = unassignedComps + [comp] #This is just another way of appending to a list

        elif comp.needsElectric:
            print(f'Needs Electric: {comp.companyName}')
            try:
                powBooth = powBooths[-1] # sets powBooth equal to the last element in the powBooths list
                # powBooth (Booth object) now has a company name for the booth name associated with it. previously, companyName was "" for booth objects
                powBooth.companyName = comp.companyName 
                finBooths = finBooths + [powBooth]
                powBooths = powBooths[0:-1]
                numBoothsAssigned += 1
            except IndexError:
                print(f'WARNING: POWER BOOTH UNABLE TO BE ASSIGNED TO {comp.companyName}')
                notEnoughPower = True
                unassignedComps = unassignedComps + [comp]

        elif 'Standard Booth' in comp.boothType:
            print(f'{comp.boothType}: {comp.companyName}')
            # search for an empty standard booth
            try:
                stdBooth = stdBooths[-1] #Sets stdBooth equal to the last element in the stdBooths list (a booth object)
                stdBooth.companyName = comp.companyName #stdBooth (Booth object) now has the company name and the booth name associated with it
                finBooths = finBooths + [stdBooth]
                stdBooths = stdBooths[0:-1]
                numBoothsAssigned += 1

            # don't have any standard booths left - try to find an empty power booth we can assign a standard company to
            except IndexError:
                try:
                    powBooth = powBooths[-1]
                    powBooth.companyName = comp.companyName
                    finBooths = finBooths + [powBooth]
                    powBooths = powBooths[0:-1]
                    numBoothsAssigned += 1

                # don't have any power booths left either
                except IndexError:
                    print(f'WARNING: BOOTH UNABLE TO BE ASSIGNED TO {comp.companyName}')
                    notEnoughBooths = True
                    unassignedComps = unassignedComps + [comp]

    print('--------------------------------------------------------------')
    print('FINISHED!')
    print(f'Total number of companies found: {len(comps)}')
    print(f'Total number of booths assigned: {numBoothsAssigned}\n')

    if notEnoughPrem or notEnoughPower or notEnoughBooths:
        printUnassignedComps(unassignedComps)

    # assign the booths to the excel file
    wb = load_workbook(filename = LAYOUTFILENAME)
    sheet_ranges = wb.worksheets[0]

    # assign companies to a booth here
    for b in finBooths: 
        cell = sheet_ranges.cell(column=b.boothCol, row=b.boothRow)
        cell.value = b.companyName

    # booths are assigned, now create a list of company names and their respective booths
    
    # to create a seperate sheet, uncomment the two lines below. 
    # make sure you replace "sheet_ranges" with "list_sheet". also, change column's value to 2 on line 159
    #wb.create_sheet('List of Companies and Booths')
    #list_sheet = wb.worksheets[1]

    # juice and contrastFill are PatternFill objects that fill the cells with color
    # juice fills cells that are power booths in with yellow
    # contrastFill fills the header cells ("Company Name" and "Booth Name") in with grey 
    juice = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
    contrastFill = PatternFill(start_color='AEAAAA', end_color='AEAAAA', fill_type='solid') 

    # top, bottom, right, and left are Side objects that are defined for each side of Border
    # border is a Border object. The Border class takes in the Side class as its parameters.
    # These objects are used to give borders to cells that are filled with juice or contrastFill in the list
    top = Side(border_style='thin', color='D0CECE')
    bottom = Side(border_style='thin', color='D0CECE')
    right = Side(border_style='thin', color='D0CECE')
    left = Side(border_style='thin', color='D0CECE')
    border = Border(top=top, bottom=bottom, right=right, left=left)

    # create the two headings ('Company Name' and 'Booth Name') of the list
    # both headers are bolded and the 'Company Name' heading is wrapped
    column = 18 # column = 2 for separate sheet , column = 18 for same sheet as layout
    row = 2
    companyCell = sheet_ranges.cell(column=column, row=row, value='Company Name')
    companyCell.alignment = Alignment(wrap_text=True)
    companyCell.font = Font(bold = True)
    companyCell.fill = contrastFill
    companyCell.border = border
    sheet_ranges.column_dimensions[get_column_letter(column)].auto_size = True
    sheet_ranges.column_dimensions[get_column_letter(column)].bestFit = True

    boothCell = sheet_ranges.cell(column=column+1, row=row, value='Booth Name')
    boothCell.alignment = Alignment(wrap_text=True)
    boothCell.font = Font(bold = True)
    boothCell.fill = contrastFill
    boothCell.border = border

    # populate cells in columns of 31 rows with company names and their respective booth name
    # the company names are wrapped and the companies/booths that need power are highlighted yellow (just like in the layout)
    temp = 0
    once = True
    for x in range( (len(finBooths)//31) + 1 ): # (len(finBooths)//31) calculates how many columns are needed to output the list
        row = 3
        once = True
        for index, b in enumerate(finBooths):
            if (x > 0):
                index = temp+1
                temp = index
                if(once):
                    column += 3
                    companyCell = sheet_ranges.cell(column=column, row=row-1, value='Company Name') # populates header cells with "Company Name"
                    companyCell.alignment = Alignment(wrap_text=True)
                    companyCell.font = Font(bold = True)
                    companyCell.fill = contrastFill
                    companyCell.border = border
                    boothCell = sheet_ranges.cell(column=column+1, row=row-1, value='Booth Name') # populates header cells with "Booth Name"
                    boothCell.alignment = Alignment(wrap_text=True)
                    boothCell.font = Font(bold = True)
                    boothCell.fill = contrastFill
                    boothCell.border = border
                    sheet_ranges.column_dimensions[get_column_letter(column)].auto_size = True
                    sheet_ranges.column_dimensions[get_column_letter(column)].bestFit = True
                    once = False
            if (index == len(finBooths)):
                break
            b = finBooths[index]
            companyCell = sheet_ranges.cell(column=column, row=row) 
            companyCell.value = b.companyName # populates cells with company names
            companyCell.alignment = Alignment(wrap_text=True)    #To wrap/unwrap the company names, uncomment/comment this line
            boothCell = sheet_ranges.cell(column=column+1, row=row)
            boothCell.value = b.boothName # populates cells with booth names    
            if b.isPower: # if a company/booth needs power, its cell will be filled with yellow and its border will be outlined
                companyCell.fill = juice
                companyCell.border = border
                boothCell.fill = juice
                boothCell.border = border

            row += 1
            if (row > 31):
                temp = index
                break

    # a list of companies and their respective booths has now been created, now figure out where to save the results

    # create a results folder (if it doesn't already exist)
    if not os.path.isdir('results'):
        os.mkdir('results')

    # create the file name
    cwd = os.getcwd()
    outputFileName = LAYOUTFILENAME[0:-5] + '_RESULTS'
    extension = '.xlsx'
    outputPath = cwd + '\\results\\' + outputFileName + extension
    
    # if we already have a results file with this name, start looking for a number
    # to stick on the end of the filename to make it unique
    # e.g. My_Booth_Layout_RESULTS (1).xlsx
    count = 1
    while os.path.exists(outputPath):
        outputPath = cwd + '\\results\\' + outputFileName + f' ({count})' + extension
        count += 1
    # found a unique name - now save it
    wb.save(outputPath)
    print(f'Result saved to: {outputPath}\n')

if __name__ == '__main__':
    main()
